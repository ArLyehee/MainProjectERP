import sys
import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

# ── 스타일 정의 ───────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_FILL    = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL  = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
RED_FILL    = PatternFill("solid", fgColor="FDECEA")
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")

WHITE_BOLD  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
DARK_BOLD   = Font(name="맑은 고딕", bold=True, color="1F3864", size=10)
DARK_NORMAL = Font(name="맑은 고딕", color="1F3864", size=10)
TITLE_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=14)
SMALL_GRAY  = Font(name="맑은 고딕", color="7F7F7F", size=9)

THIN = Side(style="thin", color="B8CCE4")
MED  = Side(style="medium", color="2E75B6")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

NUM_FMT   = '#,##0'
PCT_FMT   = '0.00"%"'

def cell_style(ws, row, col, value=None, font=None, fill=None,
               alignment=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:          c.font          = font
    if fill:          c.fill          = fill
    if alignment:     c.alignment     = alignment
    if border:        c.border        = border
    if number_format: c.number_format = number_format
    return c

def draw_table_header(ws, row, cols, fill=HEADER_FILL):
    for ci, label in enumerate(cols, 1):
        cell_style(ws, row, ci, label, WHITE_BOLD, fill, CENTER, THIN_BORDER)

def auto_width(ws, min_w=10, max_w=30):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

# ── 데이터 파싱 ───────────────────────────────────────────
raw = sys.stdin.buffer.read()
df  = pd.read_csv(io.BytesIO(raw), encoding='utf-8-sig')

total_row = df[df['월'] == '합계'].copy()
df        = df[df['월'] != '합계'].copy()

for col in ['영업이익률']:
    df[col]        = df[col].astype(str).str.replace('%', '', regex=False).astype(float)
    if not total_row.empty:
        total_row[col] = total_row[col].astype(str).str.replace('%', '', regex=False).astype(float)

num_cols = ['매출액', '매출원가', '매출총이익', '판매비및관리비', '영업이익']
for col in num_cols:
    df[col]        = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    if not total_row.empty:
        total_row[col] = pd.to_numeric(total_row[col], errors='coerce').fillna(0).astype(int)

df = df[df['매출액'] > 0].copy()

year        = df['월'].iloc[0][:4] if not df.empty else str(datetime.now().year)
total_rev   = int(df['매출액'].sum())
total_cogs  = int(df['매출원가'].sum())
total_gp    = int(df['매출총이익'].sum())
total_exp   = int(df['판매비및관리비'].sum())
total_op    = int(df['영업이익'].sum())
avg_margin  = round(df['영업이익률'].mean(), 2) if not df.empty else 0
best_month  = df.loc[df['매출액'].idxmax(), '월'] if not df.empty else '-'
worst_month = df.loc[df['영업이익'].idxmin(), '월'] if not df.empty else '-'

# ── 워크북 생성 ───────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════
# 시트 1 : 재무보고서 (커버)
# ════════════════════════════════════════════════════════
ws1 = wb.create_sheet("재무보고서")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 4
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 28

# 타이틀 배너
ws1.merge_cells('B2:E3')
c = ws1['B2']
c.value      = f"재무보고서  |  {year}년도"
c.font       = TITLE_FONT
c.fill       = BLUE_FILL
c.alignment  = CENTER

ws1.row_dimensions[2].height = 22
ws1.row_dimensions[3].height = 22

# 기본 정보
info = [
    ("회사명",   "개발환기좀해 ERP"),
    ("기준연도", f"{year}년"),
    ("작성일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("작성부서", "경영기획팀"),
]
for ri, (k, v) in enumerate(info, 5):
    cell_style(ws1, ri, 2, k, DARK_BOLD,   SUB_FILL,  LEFT, THIN_BORDER)
    cell_style(ws1, ri, 3, v, DARK_NORMAL, GRAY_FILL, LEFT, THIN_BORDER)

# KPI 요약 카드
ws1.merge_cells('B10:C10')
c = ws1['B10']
c.value = "핵심 경영지표 요약"
c.font  = WHITE_BOLD
c.fill  = HEADER_FILL
c.alignment = CENTER

kpi = [
    ("총 매출액",       total_rev,  NUM_FMT),
    ("총 영업이익",     total_op,   NUM_FMT),
    ("평균 영업이익률", avg_margin, PCT_FMT),
    ("최고 매출 월",    best_month, None),
    ("최저 영업이익 월",worst_month,None),
]
for ri, (k, v, fmt) in enumerate(kpi, 11):
    fill = GREEN_FILL if isinstance(v, (int, float)) and v >= 0 else RED_FILL
    cell_style(ws1, ri, 2, k, DARK_BOLD,   SUB_FILL, LEFT, THIN_BORDER)
    cell_style(ws1, ri, 3, v, DARK_NORMAL, fill,     RIGHT if fmt else LEFT,
               THIN_BORDER, fmt)

# ════════════════════════════════════════════════════════
# 시트 2 : 연간 요약
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("연간요약")
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:F1')
c = ws2['A1']
c.value     = f"{year}년 연간 재무 요약"
c.font      = TITLE_FONT
c.fill      = BLUE_FILL
c.alignment = CENTER
ws2.row_dimensions[1].height = 30

headers = ["항목", "금액(원)", "비율"]
draw_table_header(ws2, 3, headers)

summary_data = [
    ("매출액",       total_rev,  "100.00%"),
    ("매출원가",     total_cogs, f"{total_cogs/total_rev*100:.2f}%" if total_rev else "0%"),
    ("매출총이익",   total_gp,   f"{total_gp/total_rev*100:.2f}%"  if total_rev else "0%"),
    ("판매비및관리비",total_exp, f"{total_exp/total_rev*100:.2f}%" if total_rev else "0%"),
    ("영업이익",     total_op,   f"{total_op/total_rev*100:.2f}%"  if total_rev else "0%"),
]
for ri, (label, amt, pct) in enumerate(summary_data, 4):
    fill = GRAY_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    cell_style(ws2, ri, 1, label, DARK_BOLD,   fill, LEFT,  THIN_BORDER)
    cell_style(ws2, ri, 2, amt,   DARK_NORMAL, fill, RIGHT, THIN_BORDER, NUM_FMT)
    cell_style(ws2, ri, 3, pct,   DARK_NORMAL, fill, CENTER,THIN_BORDER)

# 합계 강조행
ri = 4 + len(summary_data)
cell_style(ws2, ri, 1, "영업이익률(평균)", DARK_BOLD, TOTAL_FILL, LEFT,  MED_BORDER)
cell_style(ws2, ri, 2, avg_margin,         DARK_BOLD, TOTAL_FILL, RIGHT, MED_BORDER, PCT_FMT)
cell_style(ws2, ri, 3, "",                 None,      TOTAL_FILL, CENTER,MED_BORDER)

auto_width(ws2, 14, 30)

# ════════════════════════════════════════════════════════
# 시트 3 : 월별 손익계산서
# ════════════════════════════════════════════════════════
ws3 = wb.create_sheet("월별 손익계산서")
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:G1')
c = ws3['A1']
c.value     = f"{year}년 월별 손익계산서"
c.font      = TITLE_FONT
c.fill      = BLUE_FILL
c.alignment = CENTER
ws3.row_dimensions[1].height = 30

cols3 = ["월", "매출액", "매출원가", "매출총이익", "판매비및관리비", "영업이익", "영업이익률(%)"]
draw_table_header(ws3, 3, cols3)

for ri, (_, row) in enumerate(df.iterrows(), 4):
    op_val = row['영업이익']
    fill   = GREEN_FILL if op_val >= 0 else RED_FILL
    row_fill = GRAY_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    cell_style(ws3, ri, 1, row['월'],            DARK_BOLD,   row_fill, CENTER, THIN_BORDER)
    cell_style(ws3, ri, 2, row['매출액'],        DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws3, ri, 3, row['매출원가'],      DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws3, ri, 4, row['매출총이익'],    DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws3, ri, 5, row['판매비및관리비'],DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws3, ri, 6, op_val,               Font(name="맑은 고딕", bold=True,
               color=("2E7D32" if op_val >= 0 else "C62828"), size=10),
               fill, RIGHT, THIN_BORDER, NUM_FMT)
    cell_style(ws3, ri, 7, row['영업이익률'],   DARK_NORMAL, fill, CENTER, THIN_BORDER, PCT_FMT)

auto_width(ws3, 12, 22)

# ════════════════════════════════════════════════════════
# 시트 4 : 합계
# ════════════════════════════════════════════════════════
ws4 = wb.create_sheet("합계")
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:G1')
c = ws4['A1']
c.value     = f"{year}년 연간 합계"
c.font      = TITLE_FONT
c.fill      = BLUE_FILL
c.alignment = CENTER
ws4.row_dimensions[1].height = 30

cols4 = ["구분", "매출액", "매출원가", "매출총이익", "판매비및관리비", "영업이익", "영업이익률(%)"]
draw_table_header(ws4, 3, cols4)

totals = [
    ("연간 합계",  total_rev, total_cogs, total_gp, total_exp, total_op, avg_margin),
]
for ri, row in enumerate(totals, 4):
    fills = [TOTAL_FILL] * 7
    cell_style(ws4, ri, 1, row[0], DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER)
    cell_style(ws4, ri, 2, row[1], DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cell_style(ws4, ri, 3, row[2], DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cell_style(ws4, ri, 4, row[3], DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cell_style(ws4, ri, 5, row[4], DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    op_font = Font(name="맑은 고딕", bold=True,
                   color=("2E7D32" if row[5] >= 0 else "C62828"), size=11)
    cell_style(ws4, ri, 6, row[5], op_font,   TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cell_style(ws4, ri, 7, row[6], DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER, PCT_FMT)

# 월별 소계도 추가
draw_table_header(ws4, 6, cols4, SUB_FILL)
for ri, (_, row) in enumerate(df.iterrows(), 7):
    row_fill = GRAY_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    op_val = row['영업이익']
    cell_style(ws4, ri, 1, row['월'],            DARK_NORMAL, row_fill, CENTER, THIN_BORDER)
    cell_style(ws4, ri, 2, row['매출액'],        DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws4, ri, 3, row['매출원가'],      DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws4, ri, 4, row['매출총이익'],    DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws4, ri, 5, row['판매비및관리비'],DARK_NORMAL, row_fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cell_style(ws4, ri, 6, op_val,               Font(name="맑은 고딕",
               color=("2E7D32" if op_val >= 0 else "C62828"), size=10),
               row_fill, RIGHT, THIN_BORDER, NUM_FMT)
    cell_style(ws4, ri, 7, row['영업이익률'],   DARK_NORMAL, row_fill, CENTER, THIN_BORDER, PCT_FMT)

auto_width(ws4, 12, 22)

# ── 출력 ─────────────────────────────────────────────────
buf = io.BytesIO()
wb.save(buf)
sys.stdout.buffer.write(buf.getvalue())
