import sys
import json
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 스타일 ────────────────────────────────────────────────
BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_FILL    = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FILL  = PatternFill("solid", fgColor="D6E4F0")
GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
RED_FILL    = PatternFill("solid", fgColor="FDECEA")

WHITE_BOLD  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
DARK_BOLD   = Font(name="맑은 고딕", bold=True, color="1F3864", size=10)
DARK_NORMAL = Font(name="맑은 고딕", color="1F3864", size=10)
TITLE_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=13)

THIN  = Side(style="thin",   color="B8CCE4")
MED   = Side(style="medium", color="2E75B6")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")
NUM_FMT = '#,##0'
PCT_FMT = '0.00"%"'

def cs(ws, r, c, val=None, font=None, fill=None, align=None, border=None, nf=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:   cell.font          = font
    if fill:   cell.fill          = fill
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if nf:     cell.number_format = nf
    return cell

def header_row(ws, row, labels, fill=HEADER_FILL):
    for ci, lbl in enumerate(labels, 1):
        cs(ws, row, ci, lbl, WHITE_BOLD, fill, CENTER, THIN_BORDER)

def title_bar(ws, merge, text):
    ws.merge_cells(merge)
    c = ws[merge.split(':')[0]]
    c.value = text;  c.font = TITLE_FONT
    c.fill  = BLUE_FILL;  c.alignment = CENTER

def auto_w(ws, mn=10, mx=28):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, mn), mx)

# ── 데이터 파싱 ───────────────────────────────────────────
raw   = sys.stdin.buffer.read()
years = json.loads(raw.decode('utf-8'))   # list of {year, months:[...]}

# 연도별 DataFrame 생성
year_dfs = {}
for yd in years:
    yr  = yd['year']
    df  = pd.DataFrame(yd['months'])
    num = ['revenue','cogs','gross_profit','expenses','operating_profit']
    for col in num:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    df['operating_margin'] = pd.to_numeric(df['operating_margin'], errors='coerce').fillna(0)
    df = df[df['revenue'] > 0].copy()
    year_dfs[yr] = df

from_year = years[0]['year']
to_year   = years[-1]['year']

# ── 워크북 ───────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════
# 시트 1 : 재무보고서 (커버)
# ════════════════════════════════════════════════════════
ws1 = wb.create_sheet("재무보고서")
ws1.sheet_view.showGridLines = False
for col, w in zip('ABCDE', [4,24,26,24,26]):
    ws1.column_dimensions[col].width = w
ws1.row_dimensions[2].height = 24
ws1.row_dimensions[3].height = 24

ws1.merge_cells('B2:E3')
c = ws1['B2']
c.value = f"재무보고서  |  {from_year}년 ~ {to_year}년"
c.font = TITLE_FONT;  c.fill = BLUE_FILL;  c.alignment = CENTER

for ri, (k, v) in enumerate([
    ("회사명",   "개발환기좀해 ERP"),
    ("기준기간", f"{from_year}년 ~ {to_year}년"),
    ("작성일시", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("작성부서", "경영기획팀"),
], 5):
    cs(ws1, ri, 2, k, DARK_BOLD,   SUB_FILL,  LEFT, THIN_BORDER)
    cs(ws1, ri, 3, v, DARK_NORMAL, GRAY_FILL, LEFT, THIN_BORDER)

# 연도별 KPI 요약
ws1.merge_cells('B10:E10')
c = ws1['B10'];  c.value = "연도별 핵심 경영지표"
c.font = WHITE_BOLD;  c.fill = HEADER_FILL;  c.alignment = CENTER

kpi_cols = ["연도", "총 매출액", "총 영업이익", "영업이익률(평균)"]
header_row(ws1, 11, kpi_cols, SUB_FILL)
for ri, (yr, df) in enumerate(year_dfs.items(), 12):
    fill = GRAY_FILL if ri % 2 == 0 else WHITE_FILL
    rev = int(df['revenue'].sum())
    op  = int(df['operating_profit'].sum())
    mg  = round(df['operating_margin'].mean(), 2)
    cs(ws1, ri, 1, str(yr),  DARK_BOLD,   fill, CENTER, THIN_BORDER)
    cs(ws1, ri, 2, rev,      DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(ws1, ri, 3, op,       Font(name="맑은 고딕", color=("2E7D32" if op>=0 else "C62828"), size=10),
       fill, RIGHT, THIN_BORDER, NUM_FMT)
    cs(ws1, ri, 4, mg,       DARK_NORMAL, fill, CENTER, THIN_BORDER, PCT_FMT)

# ════════════════════════════════════════════════════════
# 시트 2 : 연간 요약 (전 연도 비교)
# ════════════════════════════════════════════════════════
ws2 = wb.create_sheet("연간요약")
ws2.sheet_view.showGridLines = False

yr_range = f"{from_year}~{to_year}"
title_bar(ws2, 'A1:G1', f"{yr_range}년 연간 재무 요약 비교")
ws2.row_dimensions[1].height = 28

cols2 = ["연도", "매출액", "매출원가", "매출총이익", "판매비및관리비", "영업이익", "영업이익률(%)"]
header_row(ws2, 3, cols2)

for ri, (yr, df) in enumerate(year_dfs.items(), 4):
    fill = GRAY_FILL if ri % 2 == 0 else WHITE_FILL
    rev  = int(df['revenue'].sum())
    cogs = int(df['cogs'].sum())
    gp   = int(df['gross_profit'].sum())
    exp  = int(df['expenses'].sum())
    op   = int(df['operating_profit'].sum())
    mg   = round(df['operating_margin'].mean(), 2)
    op_font = Font(name="맑은 고딕", bold=True,
                   color=("2E7D32" if op >= 0 else "C62828"), size=10)
    cs(ws2, ri, 1, str(yr), DARK_BOLD,   fill, CENTER, THIN_BORDER)
    cs(ws2, ri, 2, rev,     DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(ws2, ri, 3, cogs,    DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(ws2, ri, 4, gp,      DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(ws2, ri, 5, exp,     DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(ws2, ri, 6, op,      op_font,     fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(ws2, ri, 7, mg,      DARK_NORMAL, fill, CENTER, THIN_BORDER, PCT_FMT)

# 전체 합계행
all_dfs = pd.concat(year_dfs.values()) if year_dfs else pd.DataFrame()
if not all_dfs.empty:
    tot_ri = 4 + len(year_dfs)
    t_rev  = int(all_dfs['revenue'].sum())
    t_cogs = int(all_dfs['cogs'].sum())
    t_gp   = int(all_dfs['gross_profit'].sum())
    t_exp  = int(all_dfs['expenses'].sum())
    t_op   = int(all_dfs['operating_profit'].sum())
    t_mg   = round(all_dfs['operating_margin'].mean(), 2)
    op_font = Font(name="맑은 고딕", bold=True,
                   color=("2E7D32" if t_op >= 0 else "C62828"), size=10)
    cs(ws2, tot_ri, 1, "전체 합계", DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER)
    cs(ws2, tot_ri, 2, t_rev,  DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws2, tot_ri, 3, t_cogs, DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws2, tot_ri, 4, t_gp,   DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws2, tot_ri, 5, t_exp,  DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws2, tot_ri, 6, t_op,   op_font,   TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws2, tot_ri, 7, t_mg,   DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER, PCT_FMT)

auto_w(ws2)

# ════════════════════════════════════════════════════════
# 시트 3~N : 연도별 월별 손익계산서
# ════════════════════════════════════════════════════════
for yr, df in year_dfs.items():
    ws = wb.create_sheet(f"{yr}_손익계산서")
    ws.sheet_view.showGridLines = False

    title_bar(ws, 'A1:G1', f"{yr}년 월별 손익계산서")
    ws.row_dimensions[1].height = 28

    cols = ["월", "매출액", "매출원가", "매출총이익", "판매비및관리비", "영업이익", "영업이익률(%)"]
    header_row(ws, 3, cols)

    for ri, (_, row) in enumerate(df.iterrows(), 4):
        fill = GRAY_FILL if ri % 2 == 0 else WHITE_FILL
        op   = row['operating_profit']
        op_font = Font(name="맑은 고딕", bold=True,
                       color=("2E7D32" if op >= 0 else "C62828"), size=10)
        cs(ws, ri, 1, row['month'],             DARK_BOLD,   fill, CENTER, THIN_BORDER)
        cs(ws, ri, 2, row['revenue'],           DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
        cs(ws, ri, 3, row['cogs'],              DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
        cs(ws, ri, 4, row['gross_profit'],      DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
        cs(ws, ri, 5, row['expenses'],          DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
        cs(ws, ri, 6, op,                       op_font,     fill, RIGHT,  THIN_BORDER, NUM_FMT)
        cs(ws, ri, 7, row['operating_margin'],  DARK_NORMAL, fill, CENTER, THIN_BORDER, PCT_FMT)

    # 연도 소계
    tot_ri = 4 + len(df)
    op_tot = int(df['operating_profit'].sum())
    op_font = Font(name="맑은 고딕", bold=True,
                   color=("2E7D32" if op_tot >= 0 else "C62828"), size=10)
    cs(ws, tot_ri, 1, f"{yr} 합계",        DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER)
    cs(ws, tot_ri, 2, int(df['revenue'].sum()),      DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws, tot_ri, 3, int(df['cogs'].sum()),         DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws, tot_ri, 4, int(df['gross_profit'].sum()), DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws, tot_ri, 5, int(df['expenses'].sum()),     DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws, tot_ri, 6, op_tot,                        op_font,   TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(ws, tot_ri, 7, round(df['operating_margin'].mean(), 2), DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER, PCT_FMT)

    auto_w(ws)

# ════════════════════════════════════════════════════════
# 마지막 시트 : 합계 (전 연도 통합)
# ════════════════════════════════════════════════════════
wsT = wb.create_sheet("합계")
wsT.sheet_view.showGridLines = False

title_bar(wsT, 'A1:G1', f"{yr_range}년 전체 합계")
wsT.row_dimensions[1].height = 28

header_row(wsT, 3, ["연도", "매출액", "매출원가", "매출총이익", "판매비및관리비", "영업이익", "영업이익률(%)"])

for ri, (yr, df) in enumerate(year_dfs.items(), 4):
    fill = GRAY_FILL if ri % 2 == 0 else WHITE_FILL
    op   = int(df['operating_profit'].sum())
    op_font = Font(name="맑은 고딕", color=("2E7D32" if op >= 0 else "C62828"), size=10)
    cs(wsT, ri, 1, str(yr),                         DARK_NORMAL, fill, CENTER, THIN_BORDER)
    cs(wsT, ri, 2, int(df['revenue'].sum()),         DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(wsT, ri, 3, int(df['cogs'].sum()),            DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(wsT, ri, 4, int(df['gross_profit'].sum()),    DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(wsT, ri, 5, int(df['expenses'].sum()),        DARK_NORMAL, fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(wsT, ri, 6, op,                               op_font,     fill, RIGHT,  THIN_BORDER, NUM_FMT)
    cs(wsT, ri, 7, round(df['operating_margin'].mean(), 2), DARK_NORMAL, fill, CENTER, THIN_BORDER, PCT_FMT)

if not all_dfs.empty:
    tot_ri = 4 + len(year_dfs)
    op_tot = int(all_dfs['operating_profit'].sum())
    op_font = Font(name="맑은 고딕", bold=True,
                   color=("2E7D32" if op_tot >= 0 else "C62828"), size=11)
    cs(wsT, tot_ri, 1, "전체 합계",               DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER)
    cs(wsT, tot_ri, 2, int(all_dfs['revenue'].sum()),       DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(wsT, tot_ri, 3, int(all_dfs['cogs'].sum()),          DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(wsT, tot_ri, 4, int(all_dfs['gross_profit'].sum()),  DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(wsT, tot_ri, 5, int(all_dfs['expenses'].sum()),      DARK_BOLD, TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(wsT, tot_ri, 6, op_tot,                              op_font,   TOTAL_FILL, RIGHT,  MED_BORDER, NUM_FMT)
    cs(wsT, tot_ri, 7, round(all_dfs['operating_margin'].mean(), 2), DARK_BOLD, TOTAL_FILL, CENTER, MED_BORDER, PCT_FMT)

auto_w(wsT)

# ── 출력 ─────────────────────────────────────────────────
buf = io.BytesIO()
wb.save(buf)
sys.stdout.buffer.write(buf.getvalue())
